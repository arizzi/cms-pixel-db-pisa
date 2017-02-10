#!/usr/bin/python
from openpyxl import load_workbook
import re
import datetime
import os
import glob
import sys

format='''
Sensor_id: S%s
Bare_module_ID: %s
BB_company: IZM-INFN
Comment: automatic from xlsx 
Builton: %s
Module_assembly: #WaferXYID, RocPositionInBareModule
%s
type: NotReworked
center: PISA
'''


#rootdir ='cis/'
#for subdir in os.listdir(rootdir):
#	print "Processing folder: " , subdir
#	f=glob.glob("cis/%s/*.xl*"%subdir)[0]
#	print f
if True :
	f=sys.argv[1]
#"primi-5-bare-modules.xlsx"
	wb = load_workbook(filename = r'%s'%(f))
	print "Opened"
	sheets = wb.get_sheet_names()
	sheet=sys.argv[2]
	if not sheet in sheets :
		print "Cannot find",sheet
		exit(1)
#@	sheet=sheets[0]
	ws = wb.get_sheet_by_name(sheet);
	if  ws.cell("A1").value != "ROC-WAFER-ID" :
		print "bad format, A1 is not ROC-WAFER-ID"
	rocwafer = ws.cell("A2").value
	batchid = ws.cell("C7").value
	if  ws.cell("L5").value != "Bare Module ID" :
		print "bad format, L5 is not Bare Module ID"
	s=2	
	for j in xrange(5,300):
		da=ws.cell(row=j,column=10+s).value
		se=ws.cell(row=j,column=11+s).value
		bm=ws.cell(row=j,column=9+s).value
#	print bm
#		bm="B"+se
#		print bm
		if bm is not None and re.match('B[0-9]+-[0-9]+-[0-9]',bm) :
			print "New BM found",bm
			if re.match('B[0-9]+-[0-9]+-0[0-9]',bm) :
				bm=bm[:-2]+bm[-1]
				se=se[:-2]+se[-1]
				print "correcting ID to",bm,se
				
			rocs=""
			rocprint=""
			for n in xrange(j,j+16) :
				rpos=ws.cell(row=n,column=12+s).value
				roc=ws.cell(row=n,column=13+s).value
				#print roc
				if roc is not None :
				   if len(roc) == 2 :	
					roc="0"+roc
				   roc=rocwafer+"-"+roc
				else :
				   roc=""
				rocs+=roc+","
				rocprint+=roc+",%02d\n"%int(rpos)
			print "    rocs: ",rocs	
			fo = open("auto-%s.csv"%bm,"wo")   
#2015-01-27_14h15m_1422364542
			fo.write(format%(se,bm,da.strftime("%Y-%m-%d_%Hh%Mm_%s"),rocprint))
			fo.close()
os.system("tar -czvf out.tgz auto*csv")
			
