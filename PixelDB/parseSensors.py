#!/usr/bin/python
from openpyxl import load_workbook
import re
import datetime
import os
import glob
import sys
sys.path.append("/home/cmsweb/cms-pixel-db-pisa/PixelDB")
from storm import *
from PixelDB import *
import csv


def processCSV(filename) :
	unitConversion=1e9 #nA to A
	result=[]
	if os.path.isfile(filename):
		with open(filename, 'rb') as csvfile:
			reader = csv.reader(csvfile, delimiter=';', quotechar='"')
			map=[]	
			start=[]
			i=0
			IV=[]
			iv=False
			for row in reader:
				if re.match("Posten-Nr.",row[0]) :
					start.append(i)
#					print "found:",i
				if iv :
				   if row[0].isdigit() and int(row[0]) >= 0   and int(row[0]) < 300: 
					   IV[-1].append(row)	
				   else:
				     	   iv=False
	
                                if re.match("^U ",row[0]) :
				   IV.append([])
				   iv=True	

				map.append(row)
				i+=1
			for i,iv in zip(start,IV) :
				sensorID="S%s-%s-%s" % (map[i][1],map[i+1][1],map[i+1][3])
				sensorID=re.sub(" ","",sensorID)	
				print "New sensor ", sensorID		
				out_file = open("/gpfs/ddn/cms/PIXELDB/Sensors/FactoryImport/%s.IV.txt"%sensorID,"w")
				for line in iv :
#					print "V I ", float(line[0]),float(re.sub(",",".",line[1]))*unitConversion;
					out_file.write("%s %s"%(float(line[0]),float(re.sub(",",".",line[1]))*unitConversion))
				out_file.close()
				slopeIV=float(re.sub(",",".",map[i+7][1]))
				I1=abs(float(re.sub(",",".",map[i+6][1])))
				I2=abs(float(re.sub(",",".",map[i+5][1])))
				ivTemp=float(re.sub(",",".",map[i+2][1]))
				ivDate=datetime.now()
				ivOperator=map[i-3][1]
				ivComment=map[i-1][1]
				sensor=sensorID,"/gpfs/ddn/cms/PIXELDB/Sensors/FactoryImport/%s.IV.txt"%sensorID,slopeIV,I1,I2,ivTemp,ivDate,ivOperator,ivComment
				result.append(sensor)
	else:
		print "ERROR"			 
		return None

	return result


pdb = PixelDBInterface(operator="webfrontend",center="cern")
pdb.connectToDB()
rootdir ='Datenversand_ETHZ/'
for subdir in os.listdir(rootdir):
	print "Processing folder: " , subdir
	f=glob.glob("Datenversand_ETHZ/%s/*.xl*"%subdir)[0]
	wb = load_workbook(filename = r'%s'%(f))
	sheets = wb.get_sheet_names()
	sheet=sheets[0]
	ws = wb.get_sheet_by_name(sheet);
	type = ws.cell("C5").value
	comm = ws.cell("C6").value
	batchid = ws.cell("C7").value

	print "Batch, comm, type: ", batchid, comm, type

	for i in xrange(13,37):
		ngood=ws.cell(row=i,column=15).value
		if ws.cell(row=i,column=0).value > 0 and  ngood is not None:
			waferid="%s-%s"%(batchid,ws.cell(row=i,column=0).value)
			if re.match("=>",waferid) :
				waferid=re.sub("-.*=> ","-",waferid)
				print "FIXED: ", waferid
			vdepl=ws.cell(row=i,column=2).value
			print "Wafer, vdepl, ngood :", waferid,vdepl,ngood
			IVfile=ws.cell(row=i,column=11).value
			grade={}
			grade[1]=ws.cell(row=i,column=12).value
			grade[2]=ws.cell(row=i,column=13).value
			grade[3]=ws.cell(row=i,column=14).value
			result = processCSV("Datenversand_ETHZ/%s/%s"%(subdir,IVfile))
			if result : 
			  j=0	
			  for (sensorid,ivOutFile,ivSlope,I1,I2,ivTemp,ivDate,ivOperator,ivComment),gr in zip(result,grade) :
				j+=1
		               	t = pdb.insertTransfer(Transfer("CIS","ETH"))
		               	#first check the batch
        		       	b = pdb.getBatch(batchid)
				if not b:
		        	  print "Batch %s is new, inserting it..." % batchid
	         	          b = Batch(BATCH_ID=batchid, TRANSFER_ID=t.TRANSFER_ID, PRODCENTER="CIS")  
		                  if pdb.insertBatch(b) :
        		               print "OK<br>"
                		  else :
	                        	print "FAILED<br>"
	                	#then check the wafer
		                w = pdb.getWafer(waferid)
        		        if not w:
                		   print "Wafer %s is new, inserting it..." % waferid 
		                   w = Wafer(WAFER_ID=waferid, BATCH_ID=batchid, TRANSFER_ID=t.TRANSFER_ID)
        		           if pdb.insertWafer(w) :
                		        print "OK<br>"
	                	   else :
        	                	print "FAILED<br>"
		                #insert sensor
        		        s  = pdb.getSensor(sensorid)
                		if not s :
		                   s = Sensor(sensorid,t.TRANSFER_ID,type,WAFER_ID=waferid,STATUS="INSTOCK")
        		           if pdb.insertSensor(s) :
	        	             print "<br><b> Sensor %s inserted </b>" % sensorid
				     data = Data(PFNs=ivOutFile)
				     pdb.insertData(data)
				     session = Session("CIS",ivOperator.decode("latin-1"),TYPE="IV@CIS",DATE=ivDate, COMMENT=ivComment)
				     pdb.insertSession(session)
			  	     #FIXME test = Test_IV(session.SESSION_ID,sensorid,grade[j-1],data.DATA_ID,100,I1,150,I2,ivSlope,ivDate,"CIS", COMMENT="factory data import",TEMPERATURE=ivTemp)
			  	     test = Test_IV(session.SESSION_ID,sensorid,gr,data.DATA_ID,100,I1,150,I2,ivSlope,0,"CIS", COMMENT="factory data import",TEMPERATURE=ivTemp)
			             pdb.store.add(test)
				     pdb.store.commit()
				     s.LASTTEST_SENSOR_IV=test.TEST_ID
				     pdb.store.commit()
 
 
        	        	   else :
	                	     print "<br><b> Insertion of sensor %s failed! </b>" % sensorid
		                     s =None
                		else :
	                	    print "Sensor %s already exists<br>" % sensorid


